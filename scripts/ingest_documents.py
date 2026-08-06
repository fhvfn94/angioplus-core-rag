# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import sys
import uuid
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from qdrant_client import QdrantClient
from qdrant_client.models import Distance, PointStruct, VectorParams

from app.embeddings.sentence_transformer import SentenceTransformerEmbedder

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
DEFAULT_COLLECTION = "angioplus_documents"
DEFAULT_QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6333")
# Gemini API (AI Studio): prefer gemini-embedding-001; legacy text-embedding-004 may 404 on embedContent.
DEFAULT_GEMINI_EMBEDDING_MODEL = "models/gemini-embedding-001"
DEFAULT_EMBEDDING_PROVIDER = os.getenv("EMBEDDING_PROVIDER", "gemini")


@dataclass(slots=True)
class Chunk:
    id: str
    text: str
    metadata: dict


def build_embedding_text(
    file_name: str | None,
    document_type: str | None,
    section: str | None,
    page_start: int | None,
    page_end: int | None,
    text: str,
) -> str:
    parts: list[str] = []

    if file_name:
        parts.append(f"Документ: {file_name}")
    if document_type:
        parts.append(f"Тип документа: {document_type}")
    if section:
        parts.append(f"Раздел: {section}")

    if page_start is not None or page_end is not None:
        if page_start is not None and page_end is not None:
            parts.append(f"Страницы: {page_start}-{page_end}")
        elif page_start is not None:
            parts.append(f"Страницы: {page_start}")
        else:
            parts.append(f"Страницы: {page_end}")

    if parts:
        parts.append("")
    parts.append(text or "")
    return "\n".join(parts)


class SegmentExtractor:
    """Returns ordered (page_1_based, text_block) segments."""

    def extract_segments(self, path: Path) -> list[tuple[int, str]]:
        raise NotImplementedError


class PdfSegmentExtractor(SegmentExtractor):
    def extract_segments(self, path: Path) -> list[tuple[int, str]]:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Missing dependency: pypdf") from exc

        reader = PdfReader(str(path))
        segments: list[tuple[int, str]] = []
        for idx, page in enumerate(reader.pages, start=1):
            raw = page.extract_text() or ""
            normalized = raw.replace("\r", "\n")
            normalized = re.sub(r"\n{3,}", "\n\n", normalized)
            blocks = [b.strip() for b in normalized.split("\n\n") if b.strip()]
            for block in blocks:
                segments.append((idx, block))
        return segments


class DocxSegmentExtractor(SegmentExtractor):
    def extract_segments(self, path: Path) -> list[tuple[int, str]]:
        try:
            from docx import Document
        except ImportError as exc:
            raise RuntimeError("Missing dependency: python-docx") from exc

        doc = Document(str(path))
        segments: list[tuple[int, str]] = []
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if text:
                # DOCX: no stable page boundaries in MVP; treat as logical page 1.
                segments.append((1, text))
        return segments

class ExcelFaqSegmentExtractor(SegmentExtractor):
    """
    Extracts Excel FAQ rows.

    Each non-empty row becomes one segment.
    The first row is treated as headers.
    Images inside Excel are ignored in MVP.
    """

    SKIP_SHEETS = {"WpsReserved_CellImgList"}

    def extract_segments(self, path: Path) -> list[tuple[int, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Missing dependency: openpyxl") from exc

        workbook = load_workbook(filename=str(path), data_only=True)
        segments: list[tuple[int, str]] = []

        for sheet_name in workbook.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                continue

            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [self._clean_cell(value) for value in rows[0]]
            headers = [
                header if header else f"Column {idx + 1}"
                for idx, header in enumerate(headers)
            ]

            for row_index, row in enumerate(rows[1:], start=2):
                values = [self._clean_cell(value) for value in row]

                if not any(values):
                    continue

                # Skip repeated header rows like: No. | Question | Answer | Comments
                values_cf = [value.casefold() for value in values if value]

                if "question" in values_cf and "answer" in values_cf:
                    continue

                # Skip section title rows like: Flow Velocity / Operation / Features
                non_empty_values = [value for value in values if value]
                if len(non_empty_values) == 1:
                    continue

                # For this FAQ format, real Q&A usually lives in Column 2 and Column 3
                question_candidate = values[1] if len(values) > 1 else ""
                answer_candidate = values[2] if len(values) > 2 else ""

                if not question_candidate or not answer_candidate:
                    continue

                if question_candidate.casefold() == "question" or answer_candidate.casefold() == "answer":
                    continue

                row_text_parts = [
                    f"Source Type / Тип источника: FAQ",
                    f"Language / Язык: English",
                    f"Sheet / Лист: {sheet_name}",
                    f"Row / Строка: {row_index}",
                    "",
                ]

                for header, value in zip(headers, values):
                    if not value:
                        continue

                    normalized_header = self._normalize_header(header)

                    if normalized_header in {"question", "q"}:
                        row_text_parts.append(f"Question / Вопрос:\n{value}")
                    elif normalized_header in {"answer", "a"}:
                        row_text_parts.append(f"Answer / Ответ:\n{value}")
                    elif normalized_header in {"category", "topic"}:
                        row_text_parts.append(f"Category / Категория:\n{value}")
                    else:
                        row_text_parts.append(f"{header}:\n{value}")

                    row_text_parts.append("")

                text = "\n".join(row_text_parts).strip()
                if text:
                    # page number is used as row number for Excel
                    segments.append((row_index, text))

        return segments

    @staticmethod
    def _clean_cell(value) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        text = re.sub(r"\s+", " ", text)
        return text

    @staticmethod
    def _normalize_header(header: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", header.casefold())


class ExtractorFactory:
    def __init__(self) -> None:
        self._extractors: dict[str, SegmentExtractor] = {
            ".pdf": PdfSegmentExtractor(),
            ".docx": DocxSegmentExtractor(),
            ".xlsx": ExcelFaqSegmentExtractor(),
        }

    def get(self, path: Path) -> SegmentExtractor:
        ext = path.suffix.lower()
        extractor = self._extractors.get(ext)
        if extractor is None:
            raise ValueError(f"Unsupported extension: {ext}")
        return extractor


class DocumentClassifier:
    """Classification tuned for demo filenames in this project."""

    def classify_source(self, path: Path) -> str:
        name_lower = path.name.lower()
        stem_lower = path.stem.lower()
        if path.suffix.lower() == ".xlsx":
            return "FAQ"

        # Explicit demos (order: specific filenames before generic keywords)
        if "angioplus core" in name_lower and "ifu" in name_lower:
            return "IFU"
        if stem_lower == "2025july" or name_lower.startswith("2025july"):
            return "TRAINING"
        if "competitor" in name_lower:
            return "COMMERCIAL"
        if "commercial" in name_lower or "distributor" in name_lower:
            return "COMMERCIAL"

        # L1/L2 support doc (ASCII or Cyrillic filename)
        if ("l1" in stem_lower and "l2" in stem_lower) or "l1_l2" in stem_lower:
            return "L1L2"
        # Cyrillic stem may not lower() match ASCII; normalize path string
        full = path.name.casefold()
        if "l1" in full and "l2" in full:
            return "L1L2"

        if "ifu" in name_lower:
            return "IFU"
        if "training" in name_lower:
            return "TRAINING"
        return "COMMERCIAL"

    def infer_support_level(self, source: str) -> str:
        if source == "IFU":
            return "L2"
        if source == "FAQ":
            return "L2"
        if source == "TRAINING":
            return "L1"
        if source == "L1L2":
            return "L2"
        return "L1"

    def infer_risk_level(self, chunk_text: str) -> str:
        text_cf = chunk_text.casefold()

        # Russian literals as escapes for stable Windows/editor encoding (UTF-8 source also declared).
        critical_terms_ru = (
            "\u0441\u0435\u0440\u044c\u0435\u0437\u043d\u044b\u0439 \u0438\u043d\u0446\u0438\u0434\u0435\u043d\u0442",
            "\u0441\u0435\u0440\u044c\u0451\u0437\u043d\u044b\u0439 \u0438\u043d\u0446\u0438\u0434\u0435\u043d\u0442",
            "\u0430\u0432\u0430\u0440\u0438\u044f",
            "\u043a\u0440\u0438\u0442\u0438\u0447\u0435\u0441\u043a\u0438\u0439",
            "\u0441\u0431\u043e\u0439",
            "\u043e\u0442\u043a\u0430\u0437",
            "\u043d\u0435\u0438\u0441\u043f\u0440\u0430\u0432\u043d\u043e\u0441\u0442\u044c",
            "\u043f\u043e\u0442\u0435\u0440\u044f \u0434\u0430\u043d\u043d\u044b\u0445",
        )
        critical_terms_en = (
            "failure",
            "critical",
            "hazard",
            "emergency",
            "fatal",
        )

        warning_terms_ru = (
            "\u043e\u0448\u0438\u0431\u043a\u0430",
            "\u0440\u0438\u0441\u043a",
            "\u043f\u0440\u0435\u0434\u0443\u043f\u0440\u0435\u0436\u0434\u0435\u043d\u0438\u0435",
            "\u0432\u043d\u0438\u043c\u0430\u043d\u0438\u0435",
        )
        warning_terms_en = ("warning", "caution")

        def _hits(terms: tuple[str, ...]) -> bool:
            return any(term.casefold() in text_cf for term in terms)

        if _hits(critical_terms_ru) or _hits(critical_terms_en):
            return "critical"
        if _hits(warning_terms_ru) or _hits(warning_terms_en):
            return "warning"
        return "normal"

    def infer_version(self, path: Path) -> str:
        match = re.search(r"v(\d+(?:\.\d+)*)", path.stem.lower())
        if match:
            return match.group(1)
        return "unknown"


def _is_heading_line(line_raw: str) -> bool:
    """Detect IFU-style section headings: '\\d{1,3} <title>' (single line)."""
    line = line_raw.strip()
    if "..." in line or "…" in line:
        return False

    if re.search(r"\.{5,}", line):
        return False

    if not line or "\n" in line:
        return False
    if len(line) > 220:
        return False
    m = re.match(r"^(\d{1,3})\s+(.+)$", line)
    if not m:
        return False
    title = m.group(2).strip()
    # Title starts with letter (Latin or Cyrillic) to skip numeric-only noise
    if not re.match(r"^[\w\u0400-\u04FF]", title):
        return False
    if len(title) < 3:
        return False
    return True


def _segments_to_sectioned_segments(segments: list[tuple[int, str]]) -> list[tuple[int, str, str]]:
    """
    Flatten PDF/DOCX blocks into (page, body_text, section_title).
    Parses heading lines embedded in multi-line blocks.
    """
    section = "General"
    tagged: list[tuple[int, str, str]] = []

    def flush(lines: list[str], page_no: int) -> None:
        text = "\n".join(lines).strip()
        if text:
            tagged.append((page_no, text, section))

    for page_no, raw_block in segments:
        block = raw_block.replace("\r", "\n")
        lines_queue = [ln.rstrip() for ln in block.split("\n")]
        paragraph_buf: list[str] = []

        def flush_para() -> None:
            nonlocal paragraph_buf
            flush(paragraph_buf, page_no)
            paragraph_buf = []

        idx = 0
        while idx < len(lines_queue):
            ln = lines_queue[idx].strip()
            if not ln:
                flush_para()
                idx += 1
                continue
            if _is_heading_line(ln):
                flush_para()
                section = ln
                idx += 1
                continue
            paragraph_buf.append(ln.strip())
            idx += 1
        flush_para()

    return tagged

class SemanticChunker:
    def __init__(self, min_tokens: int = 120, max_tokens: int = 500, overlap_tokens: int = 50) -> None:
        if min_tokens <= 0 or max_tokens <= 0 or min_tokens > max_tokens:
            raise ValueError("Invalid chunk token limits")
        self.min_tokens = min_tokens
        self.max_tokens = max_tokens
        self.overlap_tokens = overlap_tokens

    @staticmethod
    def _estimate_tokens(text: str) -> int:
        words = len(text.split())
        return max(1, int(words / 0.75))

    @staticmethod
    def _is_toc_like(text: str) -> bool:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) < 5:
            return False
        dotted_lines = sum(1 for line in lines if "..." in line or "…" in line)
        return dotted_lines >= 3

    def split(self, segments: list[tuple[int, str]]) -> list[tuple[str, str, int, int]]:
        body_blocks = _segments_to_sectioned_segments(segments)
        if not body_blocks:
            return []

        chunks: list[tuple[str, str, int, int]] = []
        current_blocks: list[tuple[int, str, str]] = []
        current_tokens = 0
        current_section: str | None = None

        def make_chunk(blocks: list[tuple[int, str, str]]) -> tuple[str, str, int, int] | None:
            if not blocks:
                return None

            pages = [page for page, _, _ in blocks]
            section = blocks[0][2]
            text = "\n\n".join(block_text for _, block_text, _ in blocks).strip()

            if not text:
                return None

            if self._is_toc_like(text):
                return None

            return section, text, min(pages), max(pages)

        def flush_chunk() -> None:
            nonlocal current_blocks, current_tokens
            chunk = make_chunk(current_blocks)
            if chunk is not None:
                chunks.append(chunk)
            current_blocks = []
            current_tokens = 0

        for page_no, text, section in body_blocks:
            block_tokens = self._estimate_tokens(text)

            if current_section is None:
                current_section = section

            section_changed = section != current_section

            if section_changed:
                flush_chunk()
                current_section = section

            if current_blocks and current_tokens + block_tokens > self.max_tokens:
                flush_chunk()

            current_blocks.append((page_no, text, section))
            current_tokens += block_tokens

            if current_tokens >= self.max_tokens:
                flush_chunk()

        if current_blocks:
            chunk = make_chunk(current_blocks)
            if chunk is not None:
                if self._estimate_tokens(chunk[1]) < self.min_tokens and chunks and chunks[-1][0] == chunk[0]:
                    prev_section, prev_text, prev_start, prev_end = chunks[-1]
                    chunks[-1] = (
                        prev_section,
                        f"{prev_text}\n\n{chunk[1]}",
                        min(prev_start, chunk[2]),
                        max(prev_end, chunk[3]),
                    )
                else:
                    chunks.append(chunk)

        return chunks

class BaseEmbedder:
    @property
    def model_label(self) -> str:
        return "unknown"

    def embed_texts(self, texts: Sequence[str]) -> list[list[float]]:
        raise NotImplementedError


class GeminiEmbedder(BaseEmbedder):
    def __init__(
        self,
        api_key: str,
        model_name: str = DEFAULT_GEMINI_EMBEDDING_MODEL,
    ) -> None:
        if not api_key:
            raise ValueError("Gemini API key is required")

        try:
            from google import genai
            from google.genai import types
        except ImportError as exc:
            raise RuntimeError("Missing dependency: google-genai") from exc

        self._client = genai.Client(api_key=api_key)
        self._types = types
        self._model_name = model_name

    @property
    def model_label(self) -> str:
        return self._model_name

    def embed_texts(self, texts: Sequence[str]) -> list[list[float]]:
        import time

        vectors: list[list[float]] = []

        for idx, text in enumerate(texts, start=1):
            for attempt in range(5):
                try:
                    response = self._client.models.embed_content(
                        model=self._model_name,
                        contents=text,
                        config=self._types.EmbedContentConfig(
                            task_type="RETRIEVAL_DOCUMENT",
                        ),
                    )

                    if not response.embeddings:
                        raise RuntimeError("Gemini returned no embeddings")

                    embedding = response.embeddings[0]

                    if not embedding.values:
                        raise RuntimeError("Gemini returned an empty embedding")

                    vectors.append(list(embedding.values))

                    # Free tier safety throttle
                    time.sleep(0.75)
                    break

                except Exception as exc:
                    message = str(exc)

                    if (
                        "429" in message
                        or "RESOURCE_EXHAUSTED" in message
                        or "quota" in message.lower()
                    ):
                        wait_seconds = 15 * (attempt + 1)

                        logging.warning(
                            "Embedding quota hit at item %s/%s. Waiting %s seconds. Error: %s",
                            idx,
                            len(texts),
                            wait_seconds,
                            exc,
                        )

                        time.sleep(wait_seconds)
                        continue

                    raise

            else:
                raise RuntimeError(
                    f"Failed to embed text after retries at item {idx}/{len(texts)}"
                )

        return vectors

class HashEmbedder(BaseEmbedder):
    def __init__(self, dimension: int = 256) -> None:
        self.dimension = dimension

    @property
    def model_label(self) -> str:
        return f"debug-hash-{self.dimension}"

    def embed_texts(self, texts: Sequence[str]) -> list[list[float]]:
        vectors: list[list[float]] = []
        for text in texts:
            digest = hashlib.sha256(text.encode("utf-8")).digest()
            raw = list(digest)
            vec: list[float] = []
            while len(vec) < self.dimension:
                for value in raw:
                    vec.append((value / 255.0) * 2 - 1)
                    if len(vec) >= self.dimension:
                        break
            vectors.append(vec)
        return vectors


class LocalSentenceTransformerEmbedder(BaseEmbedder):
    """Adapter exposing the shared SentenceTransformerEmbedder via BaseEmbedder."""

    def __init__(self, inner: SentenceTransformerEmbedder) -> None:
        self._inner = inner

    @property
    def model_label(self) -> str:
        return self._inner.model_label

    def embed_texts(self, texts: Sequence[str]) -> list[list[float]]:
        return self._inner.embed_documents(list(texts))


class QdrantWriter:
    def __init__(self, url: str, collection_name: str, api_key: str | None = None) -> None:
        self.client = QdrantClient(url=url, api_key=api_key)
        self.collection_name = collection_name

    def delete_collection_if_exists(self) -> None:
        try:
            self.client.delete_collection(collection_name=self.collection_name)
            logging.info("Deleted Qdrant collection %r", self.collection_name)
        except Exception as exc:
            logging.debug("Delete collection %r (may not exist): %s", self.collection_name, exc)

    def ensure_collection(self, vector_size: int) -> None:
        collections = self.client.get_collections().collections
        names = {collection.name for collection in collections}
        if self.collection_name not in names:
            self.client.create_collection(
                collection_name=self.collection_name,
                vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE),
            )
            logging.info("Created Qdrant collection %r (vector_size=%s)", self.collection_name, vector_size)

    def recreate_collection(self, vector_size: int) -> None:
        self.delete_collection_if_exists()
        self.client.create_collection(
            collection_name=self.collection_name,
            vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE),
        )
        logging.info("Recreated Qdrant collection %r (vector_size=%s)", self.collection_name, vector_size)

    def upsert(self, payload_chunks: Sequence[Chunk], vectors: Sequence[Sequence[float]], *, vector_size: int) -> None:
        if not payload_chunks:
            return
        self.ensure_collection(vector_size=len(vectors[0]))
        points = [
            PointStruct(id=chunk.id, vector=list(vector), payload={"text": chunk.text, **chunk.metadata})
            for chunk, vector in zip(payload_chunks, vectors, strict=True)
        ]
        self.client.upsert(collection_name=self.collection_name, points=points, wait=True)


class IngestionPipeline:
    def __init__(
        self,
        extractor_factory: ExtractorFactory,
        classifier: DocumentClassifier,
        chunker: SemanticChunker,
        embedder: BaseEmbedder,
        writer: QdrantWriter,
        *,
        recreate_collection: bool,
    ) -> None:
        self.extractor_factory = extractor_factory
        self.classifier = classifier
        self.chunker = chunker
        self.embedder = embedder
        self.writer = writer
        self.recreate_collection = recreate_collection

    def _load_segments(self, path: Path) -> tuple[list[tuple[int, str]], str, str]:
        extractor = self.extractor_factory.get(path)
        segments = extractor.extract_segments(path)
        source = self.classifier.classify_source(path)
        version = self.classifier.infer_version(path)
        return segments, source, version

    @staticmethod
    def _infer_topic(section: str, chunk_text: str) -> str:
        if section and section != "General":
            return section[:200]
        first_sentence = re.split(r"[.!?]\s+", chunk_text.strip())[0]
        return first_sentence[:120] if first_sentence else "General"

    def _build_chunks(
        self,
        path: Path,
        segments: list[tuple[int, str]],
        source: str,
        version: str,
    ) -> list[Chunk]:
        support_level = self.classifier.infer_support_level(source)
        output: list[Chunk] = []

        # Excel FAQ: one row = one chunk
        if path.suffix.lower() == ".xlsx":
            for row_number, row_text in segments:
                sheet_match = re.search(r"Sheet / Лист:\s*(.+)", row_text)
                question_match = re.search(
                    r"Question / Вопрос:\s*(.+?)(?:\n\n|Answer / Ответ:|$)",
                    row_text,
                    flags=re.DOTALL,
                )

                sheet_name = sheet_match.group(1).strip() if sheet_match else "Unknown"
                question = question_match.group(1).strip() if question_match else ""

                metadata = {
                    "source": "FAQ",
                    "section": sheet_name,
                    "topic": question[:200] if question else sheet_name,
                    "support_level": support_level,
                    "risk_level": self.classifier.infer_risk_level(row_text),
                    "version": version,
                    "page_start": row_number,
                    "page_end": row_number,
                    "row_number": row_number,
                    "sheet_name": sheet_name,
                    "file_name": path.name,
                    "file_path": str(path.resolve()),
                    "language": "en",
                    "document_type": "faq",
                }

                output.append(
                    Chunk(
                        id=str(uuid.uuid4()),
                        text=row_text,
                        metadata=metadata,
                    )
                )

            return output

        # PDF / DOCX: normal semantic chunking
        chunk_items = self.chunker.split(segments)

        for section, chunk_text, page_start, page_end in chunk_items:
            topic = self._infer_topic(section, chunk_text)
            metadata = {
                "source": source,
                "section": section,
                "topic": topic,
                "support_level": support_level,
                "risk_level": self.classifier.infer_risk_level(chunk_text),
                "version": version,
                "page_start": page_start,
                "page_end": page_end,
                "file_name": path.name,
                "file_path": str(path.resolve()),
                "language": "ru",
                "document_type": "document",
            }
            output.append(Chunk(id=str(uuid.uuid4()), text=chunk_text, metadata=metadata))

        return output

    def run(self, input_dir: Path, recursive: bool = True) -> dict[str, int]:
        paths = self._collect_documents(input_dir=input_dir, recursive=recursive)
        if not paths:
            logging.warning("No supported documents found in %s", input_dir)
            return {"documents": 0, "chunks": 0}

        all_chunks: list[Chunk] = []
        doc_profiles: dict[str, dict[str, int | str]] = {}
        for path in paths:
            segments, source, version = self._load_segments(path)
            char_count = sum(len(t) for _, t in segments)

            chunk_items_preview = self.chunker.split(segments)

            if not segments:
                logging.warning("Skipping empty document: %s", path)
                continue

            chunks = self._build_chunks(path, segments, source, version)
            all_chunks.extend(chunks)
            key = str(path.resolve())
            doc_profiles[key] = {
                "name": path.name,
                "source": source,
                "chars": char_count,
                "chunks": len(chunk_items_preview),
            }

        if not all_chunks:
            logging.warning("No chunks generated.")
            return {"documents": len(paths), "chunks": 0}

        vectors = self.embedder.embed_texts(
            [
                build_embedding_text(
                    file_name=chunk.metadata.get("file_name"),
                    document_type=chunk.metadata.get("document_type"),
                    section=chunk.metadata.get("section"),
                    page_start=chunk.metadata.get("page_start"),
                    page_end=chunk.metadata.get("page_end"),
                    text=chunk.text,
                )
                for chunk in all_chunks
            ]
        )
        vec_size = len(vectors[0]) if vectors else 0

        embedding_model_label = self.embedder.model_label
        counts_by_index: defaultdict[str, int] = defaultdict(int)
        for chunk in all_chunks:
            counts_by_index[chunk.metadata["file_path"]] += 1

        for fp, stats in sorted(doc_profiles.items(), key=lambda x: str(x[1]["name"])):
            persisted = counts_by_index.get(fp, 0)
            logging.info(
                "Document=%r source=%s chars=%s chunks=%s embedding_model=%r vector_size=%s "
                "(points_upserted=%s)",
                stats["name"],
                stats["source"],
                stats["chars"],
                stats["chunks"],
                embedding_model_label,
                vec_size,
                persisted,
            )

        if self.recreate_collection:
            self.writer.recreate_collection(vector_size=vec_size)
        else:
            self.writer.ensure_collection(vector_size=vec_size)

        self.writer.upsert(payload_chunks=all_chunks, vectors=vectors, vector_size=vec_size)

        return {"documents": len(paths), "chunks": len(all_chunks)}

    @staticmethod
    def _collect_documents(input_dir: Path, recursive: bool) -> list[Path]:
        if not input_dir.exists():
            raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

        iterator: Iterable[Path]
        if recursive:
            iterator = input_dir.rglob("*")
        else:
            iterator = input_dir.glob("*")
        return sorted(
            [path for path in iterator if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS],
        )


def build_embedder(
    *,
    debug_hash_embeddings: bool,
    embedding_provider: str,
    gemini_api_key: str | None,
    gemini_model: str,
) -> BaseEmbedder:
    if debug_hash_embeddings:
        logging.warning("Using HashEmbedder (debug only); not suitable for real retrieval.")
        return HashEmbedder()

    provider = (embedding_provider or os.getenv("EMBEDDING_PROVIDER") or DEFAULT_EMBEDDING_PROVIDER).strip().lower()

    if provider == "sentence_transformers":
        logging.info("Using SentenceTransformerEmbedder (local) for document embeddings.")
        try:
            inner = SentenceTransformerEmbedder()
        except RuntimeError as exc:
            logging.error("Failed to initialize SentenceTransformerEmbedder: %s", exc)
            sys.exit(2)
        return LocalSentenceTransformerEmbedder(inner)

    if provider != "gemini":
        logging.error(
            "Unknown embedding provider %r. Supported: gemini, sentence_transformers.",
            provider,
        )
        sys.exit(2)

    api_key = (gemini_api_key or os.getenv("GEMINI_API_KEY") or "").strip()
    if not api_key:
        logging.error(
            "Gemini API key is required for ingestion. Set GEMINI_API_KEY or pass --gemini-api-key. "
            "For local debugging only, use --debug-hash-embeddings. "
            "For local embeddings, set EMBEDDING_PROVIDER=sentence_transformers.",
        )
        sys.exit(2)

    try:
        return GeminiEmbedder(api_key=api_key, model_name=gemini_model)
    except Exception as exc:
        logging.error("Failed to initialize Gemini embedder: %s", exc)
        sys.exit(2)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest PDF/DOCX documents into Qdrant for AngioPlus RAG.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data/training_materials"),
        help="Path to documents directory (PDF/DOCX for RAG).",
    )
    parser.add_argument("--qdrant-url", default=DEFAULT_QDRANT_URL, help="Qdrant URL.")
    parser.add_argument("--qdrant-api-key", default=os.getenv("QDRANT_API_KEY"), help="Qdrant API key.")
    parser.add_argument("--collection", default=DEFAULT_COLLECTION, help="Qdrant collection name.")
    parser.add_argument(
        "--recreate-collection",
        action="store_true",
        help="Delete and recreate Qdrant collection before upsert.",
    )
    parser.add_argument(
        "--debug-hash-embeddings",
        action="store_true",
        help="Use deterministic hash vectors for demo/debug only.",
    )
    parser.add_argument("--gemini-api-key", default=os.getenv("GEMINI_API_KEY"), help="Gemini API key.")
    parser.add_argument(
        "--gemini-embedding-model",
        default=os.getenv("GEMINI_EMBEDDING_MODEL", DEFAULT_GEMINI_EMBEDDING_MODEL),
        help="Gemini embedding model id (e.g. models/gemini-embedding-001).",
    )
    parser.add_argument(
        "--embedding-provider",
        default=os.getenv("EMBEDDING_PROVIDER", DEFAULT_EMBEDDING_PROVIDER),
        help="Embedding provider: gemini or sentence_transformers.",
    )
    parser.add_argument(
        "--recursive",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Recursively scan input-dir (use --no-recursive to disable).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

    embedder = build_embedder(
        debug_hash_embeddings=args.debug_hash_embeddings,
        embedding_provider=args.embedding_provider,
        gemini_api_key=args.gemini_api_key,
        gemini_model=args.gemini_embedding_model,
    )

    pipeline = IngestionPipeline(
        extractor_factory=ExtractorFactory(),
        classifier=DocumentClassifier(),
        chunker=SemanticChunker(min_tokens=120, max_tokens=500, overlap_tokens=50),
        embedder=embedder,
        writer=QdrantWriter(url=args.qdrant_url, collection_name=args.collection, api_key=args.qdrant_api_key),
        recreate_collection=args.recreate_collection,
    )
    result = pipeline.run(input_dir=args.input_dir, recursive=args.recursive)
    logging.info("Ingestion done. Documents: %s, chunks: %s", result["documents"], result["chunks"])


if __name__ == "__main__":
    main()
