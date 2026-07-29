# -*- coding: utf-8 -*-
"""Unit tests for the ingestion building blocks in ``scripts/ingest_documents.py``."""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest

from scripts import ingest_documents as ingest


# ---------------------------------------------------------------- extractors


def test_segment_extractor_is_abstract():
    with pytest.raises(NotImplementedError):
        ingest.SegmentExtractor().extract_segments(Path("doc.pdf"))


def test_extractor_factory_maps_supported_extensions():
    factory = ingest.ExtractorFactory()

    assert isinstance(factory.get(Path("a.PDF")), ingest.PdfSegmentExtractor)
    assert isinstance(factory.get(Path("a.docx")), ingest.DocxSegmentExtractor)
    assert isinstance(factory.get(Path("a.xlsx")), ingest.ExcelFaqSegmentExtractor)


def test_extractor_factory_rejects_unknown_extension():
    with pytest.raises(ValueError, match="Unsupported extension: .txt"):
        ingest.ExtractorFactory().get(Path("a.txt"))


def test_pdf_extractor_splits_blocks_per_page(monkeypatch):
    class FakePage:
        def __init__(self, text):
            self._text = text

        def extract_text(self):
            return self._text

    class FakeReader:
        def __init__(self, path):
            self.pages = [
                FakePage("Первый блок\r\n\n\n\nВторой блок"),
                FakePage(None),
            ]

    import pypdf

    monkeypatch.setattr(pypdf, "PdfReader", FakeReader)

    segments = ingest.PdfSegmentExtractor().extract_segments(Path("doc.pdf"))

    assert segments == [(1, "Первый блок"), (1, "Второй блок")]


def test_docx_extractor_skips_empty_paragraphs(monkeypatch):
    def fake_document(path):
        return SimpleNamespace(
            paragraphs=[
                SimpleNamespace(text="  Заголовок  "),
                SimpleNamespace(text="   "),
                SimpleNamespace(text="Текст"),
            ]
        )

    import docx

    monkeypatch.setattr(docx, "Document", fake_document)

    segments = ingest.DocxSegmentExtractor().extract_segments(Path("doc.docx"))

    assert segments == [(1, "Заголовок"), (1, "Текст")]


def _write_faq_workbook(path: Path, rows: list[list], sheet_title: str = "FAQ") -> Path:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def test_excel_extractor_builds_one_segment_per_row(tmp_path):
    path = _write_faq_workbook(
        tmp_path / "faq.xlsx",
        [
            ["No.", "Question", "Answer", "Comments"],
            [1, "How to install?", "Run the installer", "internal"],
            [None, None, None, None],
            ["Flow Velocity", None, None, None],
            [2, "Question", "Answer", None],
            [3, "Only question", None, None],
        ],
    )

    segments = ingest.ExcelFaqSegmentExtractor().extract_segments(path)

    assert len(segments) == 1
    row_number, text = segments[0]
    assert row_number == 2
    assert "Sheet / Лист: FAQ" in text
    assert "Question / Вопрос:\nHow to install?" in text
    assert "Answer / Ответ:\nRun the installer" in text
    assert "Comments:\ninternal" in text


def test_excel_extractor_maps_category_and_unknown_headers(tmp_path):
    path = _write_faq_workbook(
        tmp_path / "faq.xlsx",
        [
            ["Topic", "Q", "A", None],
            ["Flow", "What is FFR?", "A ratio", "extra"],
        ],
    )

    segments = ingest.ExcelFaqSegmentExtractor().extract_segments(path)

    text = segments[0][1]
    assert "Category / Категория:\nFlow" in text
    assert "Question / Вопрос:\nWhat is FFR?" in text
    assert "Column 4:\nextra" in text


def test_excel_extractor_skips_reserved_sheets(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "faq.xlsx"
    workbook = Workbook()
    skipped = workbook.active
    skipped.title = "WpsReserved_CellImgList"
    skipped.append(["No.", "Question", "Answer"])
    skipped.append([1, "hidden", "hidden"])
    empty = workbook.create_sheet("Empty")
    assert empty.max_row == 1
    workbook.save(path)

    assert ingest.ExcelFaqSegmentExtractor().extract_segments(path) == []


@pytest.mark.parametrize(
    "value, expected",
    [(None, ""), ("  a   b  ", "a b"), (12, "12")],
)
def test_clean_cell(value, expected):
    assert ingest.ExcelFaqSegmentExtractor._clean_cell(value) == expected


@pytest.mark.parametrize(
    "header, expected",
    [("Question / Вопрос", "question"), ("A", "a"), ("Column 4", "column4")],
)
def test_normalize_header(header, expected):
    assert ingest.ExcelFaqSegmentExtractor._normalize_header(header) == expected


# --------------------------------------------------------------- classifier


@pytest.mark.parametrize(
    "file_name, expected",
    [
        ("FAQ_list.xlsx", "FAQ"),
        ("AngioPlus Core v2.5 IFU_RU_Final.pdf", "IFU"),
        ("2025July_training_deck.pdf", "TRAINING"),
        ("competitor_overview.pdf", "COMMERCIAL"),
        ("distributor_pricing.docx", "COMMERCIAL"),
        ("support_l1_l2_processes.docx", "L1L2"),
        ("Вопросы L1 и L2.docx", "L1L2"),
        ("ifu_supplement.pdf", "IFU"),
        ("training_notes.pdf", "TRAINING"),
        ("random_document.pdf", "COMMERCIAL"),
    ],
)
def test_classify_source(file_name, expected):
    assert ingest.DocumentClassifier().classify_source(Path(file_name)) == expected


@pytest.mark.parametrize(
    "source, expected",
    [
        ("IFU", "L2"),
        ("FAQ", "L2"),
        ("L1L2", "L2"),
        ("TRAINING", "L1"),
        ("COMMERCIAL", "L1"),
    ],
)
def test_infer_support_level(source, expected):
    assert ingest.DocumentClassifier().infer_support_level(source) == expected


@pytest.mark.parametrize(
    "text, expected",
    [
        ("Произошёл критический сбой", "critical"),
        ("Critical HAZARD detected", "critical"),
        ("Внимание: возможен риск", "warning"),
        ("Warning: check the cable", "warning"),
        ("Обычное описание функции", "normal"),
    ],
)
def test_infer_risk_level(text, expected):
    assert ingest.DocumentClassifier().infer_risk_level(text) == expected


@pytest.mark.parametrize(
    "file_name, expected",
    [
        ("AngioPlus Core v2.5 IFU.pdf", "2.5"),
        ("manual_v3.pdf", "3"),
        ("manual.pdf", "unknown"),
    ],
)
def test_infer_version(file_name, expected):
    assert ingest.DocumentClassifier().infer_version(Path(file_name)) == expected


# ------------------------------------------------------------------ chunking


@pytest.mark.parametrize(
    "line, expected",
    [
        ("11 Инсталляция ПО", True),
        ("7 Installation", True),
        ("11 Инсталляция ПО ....... 9", False),
        ("Содержание …", False),
        ("", False),
        ("1 ab", False),
        ("1 " + "x" * 300, False),
        ("plain paragraph", False),
        ("12 34", False),
    ],
)
def test_is_heading_line(line, expected):
    assert ingest._is_heading_line(line) is expected


def test_segments_to_sectioned_segments_tracks_headings():
    segments = [
        (1, "1 Введение\nПервый абзац\n\nВторой абзац"),
        (2, "2 Установка\nТретий абзац"),
    ]

    tagged = ingest._segments_to_sectioned_segments(segments)

    assert tagged == [
        (1, "Первый абзац", "1 Введение"),
        (1, "Второй абзац", "1 Введение"),
        (2, "Третий абзац", "2 Установка"),
    ]


def test_segments_to_sectioned_segments_defaults_to_general():
    tagged = ingest._segments_to_sectioned_segments([(3, "Просто текст")])

    assert tagged == [(3, "Просто текст", "General")]


@pytest.mark.parametrize(
    "min_tokens, max_tokens",
    [(0, 10), (10, 0), (20, 10)],
)
def test_semantic_chunker_validates_limits(min_tokens, max_tokens):
    with pytest.raises(ValueError, match="Invalid chunk token limits"):
        ingest.SemanticChunker(min_tokens=min_tokens, max_tokens=max_tokens)


def test_estimate_tokens_is_at_least_one():
    assert ingest.SemanticChunker._estimate_tokens("") == 1
    assert ingest.SemanticChunker._estimate_tokens("три слова тут") == 4


def test_is_toc_like_detects_dotted_lines():
    toc = "\n".join(
        [
            "1 Введение ....... 1",
            "2 Установка ....... 2",
            "3 Настройка ....... 3",
            "4 Работа",
            "5 Сервис",
        ]
    )

    assert ingest.SemanticChunker._is_toc_like(toc) is True
    assert ingest.SemanticChunker._is_toc_like("Короткий текст") is False


def test_split_returns_empty_for_no_segments():
    assert ingest.SemanticChunker().split([]) == []


def test_split_flushes_on_section_change():
    chunker = ingest.SemanticChunker(min_tokens=1, max_tokens=1000)
    segments = [
        (1, "1 Введение\nТекст первого раздела"),
        (2, "2 Установка\nТекст второго раздела"),
    ]

    chunks = chunker.split(segments)

    assert [chunk[0] for chunk in chunks] == ["1 Введение", "2 Установка"]
    assert chunks[0][1] == "Текст первого раздела"
    assert chunks[0][2:] == (1, 1)


def test_split_respects_max_tokens():
    chunker = ingest.SemanticChunker(min_tokens=1, max_tokens=10)
    long_block = " ".join(f"слово{i}" for i in range(20))
    segments = [(1, long_block), (2, long_block)]

    chunks = chunker.split(segments)

    assert len(chunks) == 2
    assert chunks[0][2] == 1 and chunks[1][2] == 2


def test_split_drops_table_of_contents_chunks():
    chunker = ingest.SemanticChunker(min_tokens=1, max_tokens=1000)
    toc_lines = "\n\n".join(f"Раздел {i} ....... {i}" for i in range(1, 7))

    assert chunker.split([(1, toc_lines)]) == []


def test_split_merges_small_tail_chunk_into_previous():
    chunker = ingest.SemanticChunker(min_tokens=40, max_tokens=40)
    long_block = " ".join(f"слово{i}" for i in range(30))
    segments = [(1, long_block), (2, "короткий хвост")]

    chunks = chunker.split(segments)

    assert len(chunks) == 1
    assert chunks[0][1].endswith("короткий хвост")
    assert chunks[0][2:] == (1, 2)


# ----------------------------------------------------------------- embedders


def test_base_embedder_contract():
    embedder = ingest.BaseEmbedder()

    assert embedder.model_label == "unknown"
    with pytest.raises(NotImplementedError):
        embedder.embed_texts(["text"])


def test_hash_embedder_is_deterministic_and_bounded():
    embedder = ingest.HashEmbedder(dimension=8)

    first = embedder.embed_texts(["текст", "другой"])
    second = embedder.embed_texts(["текст", "другой"])

    assert embedder.model_label == "debug-hash-8"
    assert first == second
    assert first[0] != first[1]
    assert all(len(vector) == 8 for vector in first)
    assert all(-1.0 <= value <= 1.0 for value in first[0])


def test_hash_embedder_pads_dimension_beyond_digest_size():
    vector = ingest.HashEmbedder(dimension=100).embed_texts(["текст"])[0]

    assert len(vector) == 100


@pytest.fixture()
def fake_generativeai(monkeypatch):
    import google.generativeai as genai

    state: dict = {"calls": [], "errors": []}

    monkeypatch.setattr(genai, "configure", lambda api_key: state.update(key=api_key))

    def embed_content(model, content, task_type):
        state["calls"].append({"model": model, "content": content})
        if state["errors"]:
            raise state["errors"].pop(0)
        return {"embedding": [0.1, 0.2]}

    monkeypatch.setattr(genai, "embed_content", embed_content)
    return state


@pytest.fixture(autouse=True)
def no_sleep(monkeypatch):
    import time

    monkeypatch.setattr(time, "sleep", lambda seconds: None)


def test_gemini_embedder_requires_api_key():
    with pytest.raises(ValueError, match="Gemini API key is required"):
        ingest.GeminiEmbedder(api_key="")


def test_gemini_embedder_embeds_texts(fake_generativeai):
    embedder = ingest.GeminiEmbedder(api_key="key", model_name="models/test")

    vectors = embedder.embed_texts(["один", "два"])

    assert vectors == [[0.1, 0.2], [0.1, 0.2]]
    assert embedder.model_label == "models/test"
    assert fake_generativeai["key"] == "key"
    assert [call["content"] for call in fake_generativeai["calls"]] == ["один", "два"]


def test_gemini_embedder_retries_on_quota_errors(fake_generativeai):
    fake_generativeai["errors"] = [RuntimeError("429 RESOURCE_EXHAUSTED")]
    embedder = ingest.GeminiEmbedder(api_key="key")

    assert embedder.embed_texts(["один"]) == [[0.1, 0.2]]
    assert len(fake_generativeai["calls"]) == 2


def test_gemini_embedder_raises_after_exhausting_retries(fake_generativeai):
    fake_generativeai["errors"] = [RuntimeError("quota exceeded") for _ in range(5)]
    embedder = ingest.GeminiEmbedder(api_key="key")

    with pytest.raises(RuntimeError, match="Failed to embed text after retries"):
        embedder.embed_texts(["один"])


def test_gemini_embedder_propagates_other_errors(fake_generativeai):
    fake_generativeai["errors"] = [RuntimeError("400 invalid model")]
    embedder = ingest.GeminiEmbedder(api_key="key")

    with pytest.raises(RuntimeError, match="400 invalid model"):
        embedder.embed_texts(["один"])


def test_build_embedder_returns_hash_embedder_in_debug_mode():
    embedder = ingest.build_embedder(
        debug_hash_embeddings=True,
        gemini_api_key=None,
        gemini_model="models/test",
    )

    assert isinstance(embedder, ingest.HashEmbedder)


def test_build_embedder_uses_env_api_key(monkeypatch, fake_generativeai):
    monkeypatch.setenv("GEMINI_API_KEY", "env-key")

    embedder = ingest.build_embedder(
        debug_hash_embeddings=False,
        gemini_api_key=None,
        gemini_model="models/test",
    )

    assert isinstance(embedder, ingest.GeminiEmbedder)
    assert fake_generativeai["key"] == "env-key"


def test_build_embedder_exits_without_api_key(monkeypatch):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    with pytest.raises(SystemExit) as exc_info:
        ingest.build_embedder(
            debug_hash_embeddings=False,
            gemini_api_key=None,
            gemini_model="models/test",
        )

    assert exc_info.value.code == 2


def test_build_embedder_exits_when_initialization_fails(monkeypatch):
    def boom(api_key, model_name):
        raise RuntimeError("no dependency")

    monkeypatch.setattr(ingest, "GeminiEmbedder", boom)

    with pytest.raises(SystemExit) as exc_info:
        ingest.build_embedder(
            debug_hash_embeddings=False,
            gemini_api_key="key",
            gemini_model="models/test",
        )

    assert exc_info.value.code == 2
